#!/usr/bin/env python3
"""Parser: WK 2026 invulformulieren (xlsx) -> deelnemers.json.

Leest alle .xlsx-bestanden in een map met het vaste format van het
Waterproef WK 2026-invulformulier en schrijft een JSON-dataset plus
een validatieverslag.

Gebruik:
    python3 parser.py <formulieren-map> <output.json>
"""
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

# Vaste celposities (zie PRD section 8)
CEL_NAAM = "H2"
RIJ_GROEPSWEDSTRIJDEN = range(25, 38)          # B=datum E=thuis G=uit H/J=score
GROEPSBLOKKEN = [                              # (headerrij, [(landkolom, plaatskolom), ...])
    (40, [("B", "D"), ("E", "F"), ("G", "H"), ("I", "L")]),
    (46, [("B", "D"), ("E", "F"), ("G", "H"), ("I", "L")]),
    (52, [("B", "D"), ("E", "F"), ("G", "H"), ("I", "L")]),
]
RIJ_ZESTIENDE = range(60, 76)                  # A=nr 1-16,  H=winnaar
RIJ_ACHTSTE = range(79, 87)                    # A=nr 17-24, H=winnaar
RIJ_KWART = range(90, 94)                      # A=nr 25-28, H=winnaar
RIJ_HALVE = range(97, 99)                      # A=nr 29-30, H=winnaar
CEL_KAMPIOEN = "H101"
CEL_TWEEDE = "H102"
BONUSCELLEN = {
    "nl_doelpunten_voor": "F104",
    "nl_doelpunten_tegen": "F105",
    "nl_geel": "F106",
    "toernooi_doelpunten": "F107",
    "toernooi_rood": "F108",
    "topscorer": "F110",
}
# Groepen-sectie (canonieke landenlijst uit het formulier zelf)
GROEPEN_HEADERRIJEN = [5, 11, 17]
GROEPEN_KOLOMMEN = ["B", "E", "G", "I"]


def norm(s):
    """Normaliseer een landnaam voor vergelijking (accenten, spaties, kast)."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().casefold()


def lees_canonieke_landen(ws):
    """Lees de 12 groepen (A-L) met elk 4 landen uit de Groepen-sectie."""
    groepen = {}
    for headerrij in GROEPEN_HEADERRIJEN:
        for kol in GROEPEN_KOLOMMEN:
            letter = ws[f"{kol}{headerrij}"].value
            if letter is None:
                continue
            landen = [ws[f"{kol}{headerrij + i}"].value for i in range(1, 5)]
            groepen[str(letter).strip()] = [str(l).strip() for l in landen if l]
    return groepen


class Validatie:
    def __init__(self, bestand):
        self.bestand = bestand
        self.fouten = []
        self.waarschuwingen = []

    def fout(self, msg):
        self.fouten.append(msg)

    def warn(self, msg):
        self.waarschuwingen.append(msg)

    def als_dict(self):
        return {"bestand": self.bestand, "fouten": self.fouten,
                "waarschuwingen": self.waarschuwingen}


def check_land(naam, canoniek_norm, vald, context):
    """Controleer of een landnaam op de canonieke lijst staat."""
    if naam is None or str(naam).strip() == "":
        vald.fout(f"{context}: leeg")
        return None
    s = str(naam).strip()
    if norm(s) not in canoniek_norm:
        vald.warn(f"{context}: onbekende landnaam '{s}'")
        return s
    return canoniek_norm[norm(s)]


def parse_formulier(pad):
    wb = openpyxl.load_workbook(pad, data_only=True)
    ws = wb.worksheets[0]
    vald = Validatie(Path(pad).name)

    groepen = lees_canonieke_landen(ws)
    canoniek_norm = {norm(l): l for landen in groepen.values() for l in landen}

    naam = ws[CEL_NAAM].value
    if not naam:
        vald.fout("Naam deelnemer (H2) ontbreekt")
        naam = Path(pad).stem

    # A. Groepswedstrijden
    wedstrijden = []
    for r in RIJ_GROEPSWEDSTRIJDEN:
        thuis, uit = ws[f"E{r}"].value, ws[f"G{r}"].value
        if thuis is None and uit is None:
            continue
        datum = ws[f"B{r}"].value
        st, su = ws[f"H{r}"].value, ws[f"J{r}"].value
        if st is None or su is None:
            vald.fout(f"Groepswedstrijd rij {r} ({thuis}-{uit}): uitslag ontbreekt")
        wedstrijden.append({
            "datum": datum.strftime("%Y-%m-%d") if hasattr(datum, "strftime") else str(datum),
            "thuis": check_land(thuis, canoniek_norm, vald, f"rij {r} thuis"),
            "uit": check_land(uit, canoniek_norm, vald, f"rij {r} uit"),
            "thuis_score": st if isinstance(st, (int, float)) else None,
            "uit_score": su if isinstance(su, (int, float)) else None,
        })
    if len(wedstrijden) != 13:
        vald.warn(f"Verwacht 13 groepswedstrijden, gevonden: {len(wedstrijden)}")

    # B. Groepseindstanden -> per groep een lijst op plaats 1-4
    eindstand = {}
    for headerrij, paren in GROEPSBLOKKEN:
        for landkol, plaatskol in paren:
            letter = ws[f"{landkol}{headerrij}"].value
            if letter is None:
                continue
            letter = str(letter).strip()
            plekken = {}
            for i in range(1, 5):
                land = check_land(ws[f"{landkol}{headerrij + i}"].value,
                                  canoniek_norm, vald, f"groep {letter} rij {headerrij + i}")
                plaats = ws[f"{plaatskol}{headerrij + i}"].value
                if plaats not in (1, 2, 3, 4):
                    vald.fout(f"Groep {letter}, {land}: ongeldige plaats '{plaats}'")
                    continue
                if plaats in plekken:
                    vald.fout(f"Groep {letter}: plaats {plaats} dubbel gebruikt")
                plekken[plaats] = land
            eindstand[letter] = [plekken.get(p) for p in (1, 2, 3, 4)]
    if len(eindstand) != 12:
        vald.warn(f"Verwacht 12 groepen, gevonden: {len(eindstand)}")

    # C. Knock-outrondes: per wedstrijdnummer de voorspelde winnaar
    def lees_winnaars(rijen, context):
        winnaars = {}
        for r in rijen:
            nr = ws[f"A{r}"].value
            land = check_land(ws[f"H{r}"].value, canoniek_norm, vald, f"{context} W{nr}")
            if nr is not None:
                winnaars[str(int(nr))] = land
        return winnaars

    zestiende = lees_winnaars(RIJ_ZESTIENDE, "zestiende finale")
    achtste = lees_winnaars(RIJ_ACHTSTE, "achtste finale")
    kwart = lees_winnaars(RIJ_KWART, "kwartfinale")
    halve = lees_winnaars(RIJ_HALVE, "halve finale")

    # D. Finale
    kampioen = check_land(ws[CEL_KAMPIOEN].value, canoniek_norm, vald, "kampioen")
    tweede = check_land(ws[CEL_TWEEDE].value, canoniek_norm, vald, "verliezend finalist")

    # E. Bonusvragen
    bonus = {}
    for sleutel, cel in BONUSCELLEN.items():
        w = ws[cel].value
        if w is None:
            vald.fout(f"Bonusvraag '{sleutel}' ({cel}) leeg")
        bonus[sleutel] = str(w).strip() if sleutel == "topscorer" and w is not None else w

    return {
        "naam": str(naam).strip(),
        "bestand": Path(pad).name,
        "groepswedstrijden": wedstrijden,
        "groepseindstand": eindstand,
        "zestiende_winnaars": zestiende,
        "achtste_winnaars": achtste,
        "kwart_winnaars": kwart,
        "halve_winnaars": halve,
        "kampioen": kampioen,
        "tweede": tweede,
        "bonus": bonus,
    }, vald, groepen


def main(formmap, outpad):
    formmap = Path(formmap)
    bestanden = sorted(p for p in formmap.glob("*.xlsx") if not p.name.startswith("~"))
    if not bestanden:
        print(f"Geen xlsx-bestanden gevonden in {formmap}", file=sys.stderr)
        sys.exit(1)
    deelnemers, verslagen, groepen = [], [], None
    for pad in bestanden:
        try:
            d, vald, g = parse_formulier(pad)
            groepen = groepen or g
            deelnemers.append(d)
            verslagen.append(vald.als_dict())
        except Exception as e:  # formulier onleesbaar: rapporteren, doorgaan
            verslagen.append({"bestand": pad.name,
                              "fouten": [f"Onleesbaar: {e}"], "waarschuwingen": []})
    data = {"groepen": groepen, "deelnemers": deelnemers, "validatie": verslagen}
    Path(outpad).write_text(json.dumps(data, ensure_ascii=False, indent=2),
                            encoding="utf-8")
    n_fout = sum(1 for v in verslagen if v["fouten"])
    print(f"{len(deelnemers)} formulieren geparst -> {outpad}; "
          f"{n_fout} met fouten, zie validatieverslag in de JSON.")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
